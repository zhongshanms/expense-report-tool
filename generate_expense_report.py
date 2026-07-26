# -*- coding: utf-8 -*-
"""
费用报销 Excel 自动生成工具 - 本地版 v2.0
新增功能：
  - PDF 发票金额自动提取（pdfplumber 解析）
  - 发票文件自动重命名为纯数字金额（3位小数）
  - 重名冲突处理（自动加 0, 1, 2...）
  - 文件夹自动重命名（序号#类别+总金额元）
"""

import os
import re
import sys
import argparse
from pathlib import Path

try:
    import pdfplumber
    HAS_PDFPLUMBER = True
except ImportError:
    HAS_PDFPLUMBER = False
    print("[警告] pdfplumber 未安装，将仅从文件名提取金额。")
    print("       安装: pip install pdfplumber")

# ─── 修复 Windows 控制台中文字符编码 ───
if sys.platform == "win32":
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")


# ─── 常量 ───
QUALITY_KEYWORDS = ["质量", "检验", "QC", "QA"]

# 发票类别关键词
CATEGORY_RULES = [
    (["通行费", "高速", "公路通行", "经营租赁*通行费"], "高速通行费"),
    (["成品油", "汽油", "柴油", "加油", "加油站", "石油"], "加油费"),
    (["停车"], "停车费"),
    (["维修", "保养"], "维修费"),
    (["住宿", "酒店", "宾馆"], "住宿费"),
    (["餐饮", "餐费", "吃饭", "饭店"], "餐饮费"),
]
DEFAULT_CATEGORY = "汽车费用"


# ─── PDF 解析 ───
def extract_from_pdf(pdf_path: str) -> tuple:
    """
    从 PDF 发票中提取 (金额, 类别, 全文)
    返回: (amount_float, category_str, full_text_str)
    """
    text = ""
    amount = None
    category = DEFAULT_CATEGORY

    if not HAS_PDFPLUMBER:
        return None, category, ""

    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                t = page.extract_text()
                if t:
                    text += t + "\n"
    except Exception as e:
        return None, category, f"[解析失败] {e}"

    if not text.strip():
        return None, category, text

    # 1. 提取价税合计金额
    # 模式：(小写) ¥123.45 或 （小写）¥ 123.45 或 (小写)123.45
    patterns = [
        r'[（(]小写[）)]\s*[¥￥]\s*(\d+\.?\d*)',
        r'[（(]小写[）)]\s*(\d+\.?\d*)',
        r'价税合计.*?(\d+\.?\d+)',
        r'合\s*计[^¥]*[¥￥]\s*(\d+\.?\d+)',
    ]
    for pat in patterns:
        m = re.search(pat, text, re.DOTALL)
        if m:
            try:
                amount = float(m.group(1))
                break
            except ValueError:
                pass

    # 2. 推断费用类别
    for keywords, cat in CATEGORY_RULES:
        for kw in keywords:
            if kw in text:
                category = cat
                break
        if category != DEFAULT_CATEGORY:
            break

    return amount, category, text


def extract_amount_from_name(name: str):
    """从文件/文件夹名中提取最后一个数字作为金额（回退方案）"""
    clean = name.replace(".pdf", "").strip()
    nums = re.findall(r"\d+(?:\.\d+)?", clean)
    if nums:
        try:
            return float(nums[-1])
        except ValueError:
            pass
    return None


# ─── 金额格式化 ───
def fmt_amount(amount: float) -> str:
    """金额 → 3位小数字符串（如 316.300）"""
    return f"{amount:.3f}"


# ─── 文件夹处理 ───
def process_folder(full_folder_path: str, dry_run: bool = True) -> dict:
    """
    处理单个发票文件夹：
    1. 解析每个 PDF 提取金额
    2. 重命名 PDF 为纯数字（处理重名）
    3. 重命名文件夹为 序号#类别总金额元
    返回: {items: [...], renamed: [...], folder_new_name: str, total: float}
    """
    folder = Path(full_folder_path)
    if not folder.is_dir():
        print(f"  [跳过] 不是文件夹: {folder.name}")
        return None

    pdfs = [f for f in folder.iterdir() if f.suffix.lower() == ".pdf"]
    if not pdfs:
        print(f"  [跳过] 无PDF: {folder.name}")
        return None

    # 解析所有PDF
    pdf_data = []  # [(old_path, amount, category)]
    categories = []
    for pdf_path in sorted(pdfs):
        amount, cat, text = extract_from_pdf(str(pdf_path))
        if amount is None:
            # 回退：从文件名提取
            amount = extract_amount_from_name(pdf_path.stem)
            cat = DEFAULT_CATEGORY
            for keywords, c in CATEGORY_RULES:
                if any(kw in pdf_path.stem for kw in keywords):
                    cat = c
                    break
        if amount is not None:
            pdf_data.append((pdf_path, amount, cat))
            categories.append(cat)
        else:
            print(f"  [警告] 无法提取金额: {pdf_path.name}")

    if not pdf_data:
        return None

    total = sum(a for _, a, _ in pdf_data)

    # 确定主类别（多数投票）
    main_cat = max(set(categories), key=categories.count) if categories else DEFAULT_CATEGORY

    # 重命名PDF（处理重名冲突）
    renames = []
    used_names = {}  # amount_str -> count
    for old_path, amount, cat in pdf_data:
        base_name = fmt_amount(amount)
        if base_name in used_names:
            used_names[base_name] += 1
            new_stem = f"{base_name}-{used_names[base_name] - 1}"
        else:
            used_names[base_name] = 0
            new_stem = base_name
        new_name = new_stem + ".pdf"
        renames.append((old_path, old_path.parent / new_name))

    # 检查重名冲突
    new_names_list = [new.name for _, new in renames]
    if len(new_names_list) != len(set(new_names_list)):
        print("  [警告] 仍有重名冲突，正在重新分配...")
        name_count = {}
        final_renames = []
        for old_path, amount, cat in pdf_data:
            base = fmt_amount(amount)
            if base not in name_count:
                name_count[base] = 0
                final_renames.append((old_path, old_path.parent / f"{base}.pdf"))
            else:
                final_name = f"{base}-{name_count[base]}"
                name_count[base] += 1
                final_renames.append((old_path, old_path.parent / f"{final_name}.pdf"))
        renames = final_renames

    # 文件夹新名称
    total_str = fmt_amount(total).rstrip('0').rstrip('.')
    # 检查文件夹名中的序号前缀
    seq_match = re.match(r"^(\d+)#", folder.name)
    seq_prefix = seq_match.group(1) + "#" if seq_match else ""
    new_folder_name = f"{seq_prefix}{main_cat}{total_str}元"

    # 执行重命名
    if not dry_run:
        for old_path, new_path in renames:
            if old_path != new_path and not new_path.exists():
                old_path.rename(new_path)
                print(f"    [更名] {old_path.name} -> {new_path.name}")
            elif old_path != new_path and new_path.exists():
                print(f"    [跳过] {new_path.name} 已存在")

        # 文件夹重命名
        new_folder_path = folder.parent / new_folder_name
        if folder != new_folder_path and not new_folder_path.exists():
            folder.rename(new_folder_path)
            print(f"  [更名文件夹] {folder.name} -> {new_folder_name}")

    result = {
        "item": {
            "部门": "产品质量",
            "二级对象": "产品质量",
            "三级": "产品质量",
            "费用类别": main_cat,
            "金额": round(total, 2),
            "费用说明": f"{new_folder_name}（含{len(pdf_data)}张发票）",
        },
        "renames": [(str(o), str(n)) for o, n in renames],
        "folder_new_name": new_folder_name,
        "total": round(total, 2),
        "category": main_cat,
        "pdf_count": len(pdf_data),
    }
    return result


# ─── 主逻辑 ───
def parse_invoice_folder(folder_path: str, dry_run: bool = True) -> list:
    """
    扫描发票文件夹，自动解析并返回费用条目。
    dry_run=True 时不实际修改文件，仅供生成 Excel 和预览。
    """
    items = []
    base = Path(folder_path)

    sub_dirs = sorted([d for d in base.iterdir() if d.is_dir()])
    root_pdfs = sorted([f for f in base.iterdir() if f.is_file() and f.suffix.lower() == ".pdf"])

    # 处理子文件夹
    for sub in sub_dirs:
        result = process_folder(str(sub), dry_run=dry_run)
        if result:
            items.append(result["item"])

    # 处理根目录 PDF
    for pdf in root_pdfs:
        amount, cat, text = extract_from_pdf(str(pdf))
        if amount is None:
            amount = extract_amount_from_name(pdf.stem)
            cat = DEFAULT_CATEGORY
            for keywords, c in CATEGORY_RULES:
                if any(kw in pdf.stem for kw in keywords):
                    cat = c
                    break
        name = pdf.stem

        # 检查质量部
        dept = "产品质量"
        for kw in QUALITY_KEYWORDS:
            if kw in name:
                dept = "质量部"
                break

        items.append({
            "部门": dept,
            "二级对象": dept,
            "三级": dept,
            "费用类别": cat,
            "金额": round(amount, 2) if amount else 0.0,
            "费用说明": name,
        })

    # 排序
    def _get_seq(item):
        m = re.match(r"^(\d+)#", item["费用说明"])
        return int(m.group(1)) if m else 9999

    items.sort(key=_get_seq)

    # 编号
    for i, item in enumerate(items):
        item["序号"] = i + 1

    return items


# ─── Excel 生成（同之前的逻辑） ───
def create_excel(items: list, output_path: str):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("[错误] openpyxl 未安装。 pip install openpyxl")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "费用报销"

    # 列宽
    ws.column_dimensions["A"].width = 9
    ws.column_dimensions["B"].width = 17.3
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 11
    ws.column_dimensions["F"].width = 9
    ws.column_dimensions["G"].width = 40

    headers = ["★序号", "费用承担部门", "二级对象（店铺）", "三级（asin或其他）", "费用类别", "金额", "费用说明"]

    # 通用样式
    thin_border = Border(
        top=Side(style="thin", color="FF000000"),
        left=Side(style="thin", color="FF000000"),
        bottom=Side(style="thin", color="FF000000"),
        right=Side(style="thin", color="FF000000"),
    )
    hdr_font = Font(name="微软雅黑", size=10, bold=False, color="FF000000")
    hdr_fill = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_font = Font(name="微软雅黑", size=10, color="FF000000")
    cell_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 表头
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = thin_border
    ws.row_dimensions[1].height = 31

    # 数据行
    for r, item in enumerate(items, 2):
        row_data = [
            item.get("序号", r - 1),
            item["部门"],
            item["二级对象"],
            item["三级"],
            item["费用类别"],
            item["金额"],
            item["费用说明"],
        ]
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = cell_font
            cell.alignment = cell_align
            cell.border = thin_border
        ws.row_dimensions[r].height = 18

    # 合计行
    total_row = len(items) + 2
    total = sum(item["金额"] for item in items)
    total_data = ["", "", "", "", "合计", round(total, 2), ""]
    for c, val in enumerate(total_data, 1):
        cell = ws.cell(row=total_row, column=c, value=val)
        cell.font = Font(name="微软雅黑", size=10, bold=(c in (5, 6)), color="FF000000")
        cell.alignment = cell_align
        cell.border = thin_border
    ws.row_dimensions[total_row].height = 18

    wb.save(output_path)
    print(f"\n[完成] Excel已生成: {output_path}")
    print(f"   共 {len(items)} 条费用，合计金额: ¥{total:.2f}")


# ─── 命令行入口 ───
def main():
    parser = argparse.ArgumentParser(description="费用报销Excel自动生成工具 v2.0")
    parser.add_argument("folder", nargs="?", help="发票文件夹路径")
    parser.add_argument("-o", "--output", help="输出Excel路径")
    parser.add_argument("--dry-run", action="store_true", default=True,
                        help="仅预览，不实际重命名文件（默认）")
    parser.add_argument("--rename", action="store_true",
                        help="实际执行文件重命名")
    args = parser.parse_args()

    if args.folder:
        input_folder = args.folder
    else:
        # 默认扫描桌面
        desktop = Path.home() / "Desktop"
        candidates = sorted([d for d in desktop.iterdir()
                             if d.is_dir() and "中山" in d.name and "费用报销" in d.name],
                            reverse=True)
        if candidates:
            input_folder = str(candidates[0])
        else:
            print("未找到发票文件夹，请指定路径。")
            return

    do_rename = args.rename
    if do_rename:
        print("\n[注意] 将实际重命名文件！请确保已备份。")
        confirm = input("确认继续? (y/N): ")
        if confirm.lower() != 'y':
            print("已取消。")
            return

    items = parse_invoice_folder(input_folder, dry_run=not do_rename)

    if not items:
        print("未识别到任何费用条目。")
        return

    print(f"\n[识别] 共 {len(items)} 条费用:")
    for item in items:
        print(f"   [{item['序号']}] {item['费用说明']} -> ¥{item['金额']} ({item['费用类别']})")

    # 输出路径
    folder_path = Path(input_folder)
    parent_dir = str(folder_path.parent)
    folder_name = folder_path.name
    output_path = args.output or os.path.join(parent_dir, f"{folder_name}_生成.xlsx")

    create_excel(items, output_path)


if __name__ == "__main__":
    main()
